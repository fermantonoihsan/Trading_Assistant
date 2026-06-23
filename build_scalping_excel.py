import argparse
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart

try:
    from idx_scraper import fetch_ohlcv, fetch_stock
except Exception as exc:
    fetch_ohlcv = None
    fetch_stock = None
    OHLCV_FETCH_IMPORT_ERROR = exc
else:
    OHLCV_FETCH_IMPORT_ERROR = None

wb = Workbook()
wb.remove(wb.active)

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY      = "1B3A59"
BLUE      = "2E5E8A"
D_GREEN   = "155724"
D_RED     = "721C24"
D_AMBER   = "856404"
INPUT_BG  = "D6EAF8"    # light blue → blue text (inputs user fills)
CALC_BG   = "F2F3F4"    # gray → black text (formulas)
POS_BG    = "D5F5E3"
NEG_BG    = "FADBD8"
WARN_BG   = "FFF3CD"
WHITE     = "FFFFFF"
LTBLUE    = "EBF5FB"
LTGRAY    = "F8F9FA"
GRAY_BDR  = "BDC3C7"

# ── Helpers ───────────────────────────────────────────────────────────────────
def font(bold=False, size=10, color="000000", italic=False, underline=None):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic, underline=underline)

def fill(c):
    return PatternFill("solid", fgColor=c)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color=GRAY_BDR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_only(color=GRAY_BDR):
    return Border(bottom=Side(style="thin", color=color))

def sc(cell, bold=False, fs=10, fc="000000", bg=None, h="left", wrap=False, italic=False, num_fmt=None):
    cell.font = font(bold, fs, fc, italic)
    if bg: cell.fill = fill(bg)
    cell.alignment = align(h, "center", wrap)
    if num_fmt: cell.number_format = num_fmt

def header_row(ws, row, cols, labels, bg=NAVY, fc=WHITE, fs=10, height=22):
    ws.row_dimensions[row].height = height
    for col, lbl in zip(cols, labels):
        c = ws.cell(row=row, column=col, value=lbl)
        c.font = font(True, fs, fc)
        c.fill = fill(bg)
        c.alignment = align("center", "center")
        c.border = thin_border("AAAAAA")

def section_title(ws, row, col, text, colspan=1, bg=BLUE):
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(True, 11, WHITE)
    c.fill = fill(bg)
    c.alignment = align("left", "center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)

def set_w(ws, col_widths):
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

DEFAULT_STOCKS = ["BBRI", "BBCA", "ADRO", "TLKM", "GOTO"]

def parse_stock_codes(raw):
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        raw = " ".join(str(item) for item in raw)
    cleaned = str(raw).replace(",", " ")
    codes = []
    for part in cleaned.split():
        code = part.strip().upper()
        if code.startswith("IDX:"):
            code = code[4:]
        if code.endswith(".JK"):
            code = code[:-3]
        if code and code.replace("-", "").isalnum():
            codes.append(code)
    return list(dict.fromkeys(codes))

def resolve_stocks(cli_stocks):
    stocks = parse_stock_codes(cli_stocks)
    if stocks:
        return stocks

    prompt = (
        "Masukkan kode saham yang ingin dibuat di Scalping_Analysis.xlsx\n"
        "Contoh: BBRI BBCA ANTM atau BBRI,BBCA,ANTM\n"
        "Enter untuk default BBRI BBCA ADRO TLKM GOTO: "
    )
    stocks = parse_stock_codes(input(prompt))
    return stocks or DEFAULT_STOCKS

def clean_number(value):
    if value is None:
        return None
    try:
        if value != value:
            return None
    except Exception:
        pass
    try:
        if hasattr(value, "item"):
            value = value.item()
    except Exception:
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def build_gf_import_rows(kode, df):
    if df is None or getattr(df, "empty", False):
        return []
    rows = []
    for idx, rec in df.tail(10).iterrows():
        raw = str(idx)[:10]
        # Simpan sebagai tanggal Excel asli agar MAXIFS/INDEX-MATCH di Price_Chart jalan
        try:
            tanggal = datetime.strptime(raw, "%Y-%m-%d").date()
        except ValueError:
            tanggal = raw
        change_pct = clean_number(rec.get("Change (%)"))
        rows.append((
            tanggal,
            clean_number(rec.get("Open")),
            clean_number(rec.get("High")),
            clean_number(rec.get("Low")),
            clean_number(rec.get("Close")),
            clean_number(rec.get("Volume")),
            None if change_pct is None else change_pct / 100,
        ))
    return rows

# Cache info lengkap per saham (1x fetch_stock per kode, dipakai 3 sheet)
_STOCK_INFO_CACHE = {}
def get_stock_info(kode):
    if kode in _STOCK_INFO_CACHE:
        return _STOCK_INFO_CACHE[kode]
    info = None
    if fetch_stock is not None:
        try:
            info = fetch_stock(kode)
        except Exception:
            info = None
    _STOCK_INFO_CACHE[kode] = info
    return info

def fetch_gf_import_rows(kode):
    info = get_stock_info(kode)
    df = info.get("ohlcv") if info else None
    if df is None and fetch_ohlcv is not None:
        df = fetch_ohlcv(kode, 10)
    return build_gf_import_rows(kode, df)

def pe_row_from_info(info):
    """fetch_stock → tuple 13 nilai PE_PBV_Band (None = sel kosong)."""
    if not info:
        return None
    pe  = info.get("peBand")  or {}
    pbv = info.get("pbvBand") or {}
    return (
        clean_number(info.get("price")), clean_number(info.get("eps")), clean_number(info.get("bvps")),
        pe.get("min"), pe.get("sd_minus1"), pe.get("median"), pe.get("sd_plus1"), pe.get("max"),
        pbv.get("min"), pbv.get("sd_minus1"), pbv.get("median"), pbv.get("sd_plus1"), pbv.get("max"),
    )

def cons_row_from_info(info):
    """fetch_stock → tuple 14 nilai Consensus_EPS. Revenue Triliun→Miliar agar cocok header."""
    if not info:
        return None
    def miliar(key):
        v = info.get(key)
        return None if v is None else v * 1000
    return (
        info.get("analystBuy"), info.get("analystHold"), info.get("analystSell"),
        clean_number(info.get("targetLow")), clean_number(info.get("targetMean")), clean_number(info.get("targetHigh")),
        clean_number(info.get("epsCurrentYear")), clean_number(info.get("epsNextYear")), None,
        miliar("revCurrentYear"), miliar("revNextYear"), None,
        info.get("nextEarnings"), None,
    )

parser = argparse.ArgumentParser(description="Build Scalping_Analysis.xlsx")
parser.add_argument(
    "--stocks", nargs="*", default=None,
    help="Daftar kode saham non-interaktif (contoh: --stocks BBRI ANTM BUMI atau --stocks BBRI,ANTM). "
         "Jika tidak diisi, script akan meminta input di terminal.")
parser.add_argument(
    "--timestamp", action="store_true",
    help="Tambahkan timestamp ke nama file output agar tidak menimpa file yang sedang dibuka.")
args = parser.parse_args()
STOCKS = resolve_stocks(args.stocks)
DROPDOWN = '"' + ",".join(STOCKS) + '"'   # string DataValidation dropdown dinamis

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — GF_Import
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("GF_Import")
ws1.sheet_view.showGridLines = True
ws1.freeze_panes = "A5"

# Title
ws1.row_dimensions[1].height = 30
ws1.merge_cells("A1:H1")
c = ws1["A1"]
c.value = "📥  GOOGLE FINANCE — OHLCV IMPORT"
c.font = font(True, 14, NAVY)
c.fill = fill(LTBLUE)
c.alignment = align("center", "center")

ws1.row_dimensions[2].height = 40
ws1.merge_cells("A2:H2")
c = ws1["A2"]
c.value = (
    "Data GF_Import diisi otomatis dari idx_scraper.fetch_ohlcv untuk 10 hari terakhir "
    "per kode saham input. Jika data tidak tersedia, baris kosong tetap dibuat untuk diisi manual."
)
c.font = font(False, 9, "555555", italic=True)
c.fill = fill(WARN_BG)
c.alignment = align("left", "center", wrap=True)

# Sub-header: GF CSV standard columns
header_row(ws1, row=4,
    cols=  [1,    2,      3,     4,     5,     6,      7,      8],
    labels=["KODE","Tanggal","Open","High","Low","Close","Volume","% Change"],
    bg=NAVY, fs=10)

# Sample data rows (5 stocks × 10 rows each)
row = 5
first_data_row = row
if OHLCV_FETCH_IMPORT_ERROR:
    print("Peringatan: tidak bisa memakai idx_scraper.fetch_ohlcv:", OHLCV_FETCH_IMPORT_ERROR)
print("Mengambil OHLCV 10 hari terakhir untuk:", ", ".join(STOCKS))
for kode in STOCKS:
    # Saham tanpa sample → 10 baris input kosong untuk diisi manual
    rows = fetch_gf_import_rows(kode)
    if rows:
        print(f"  {kode}: {len(rows)} baris OHLCV")
    else:
        print(f"  {kode}: data OHLCV tidak tersedia, dibuat 10 baris kosong")
        rows = [("", None, None, None, None, None, None) for _ in range(10)]
    for tanggal, o, h, l, cl, vol, chg in rows:
        ws1.cell(row=row, column=1, value=kode)
        bcell = ws1.cell(row=row, column=2, value=tanggal)
        if not isinstance(tanggal, str):
            bcell.number_format = "yyyy-mm-dd"
        ws1.cell(row=row, column=3, value=o).number_format = '#,##0'
        ws1.cell(row=row, column=4, value=h).number_format = '#,##0'
        ws1.cell(row=row, column=5, value=l).number_format = '#,##0'
        ws1.cell(row=row, column=6, value=cl).number_format = '#,##0'
        ws1.cell(row=row, column=7, value=vol).number_format = '#,##0'
        # % Change formula vs previous close (lewati bila baris kosong)
        if chg is not None:
            ws1.cell(row=row, column=8).value = chg
            ws1.cell(row=row, column=8).number_format = "0.00%"
        elif cl is not None and row > first_data_row and ws1.cell(row=row-1, column=1).value == kode:
            ws1.cell(row=row, column=8).value = f"=(F{row}-F{row-1})/F{row-1}"
            ws1.cell(row=row, column=8).number_format = "0.00%"
        else:
            ws1.cell(row=row, column=8).value = "-"

        # Zebra + border
        row_bg = WHITE if row % 2 == 0 else LTGRAY
        for col in range(1, 9):
            c = ws1.cell(row=row, column=col)
            c.fill = fill(row_bg)
            c.border = thin_border()
            c.alignment = align("center" if col >= 3 else "left", "center")
            if col in (3,4,5,6,7):
                c.font = font(False, 10, "00008B")  # blue = hardcoded input
        row += 1

# Column widths
set_w(ws1, {1:10, 2:14, 3:12, 4:12, 5:12, 6:12, 7:16, 8:12})

# Data validation: stock code in col A
dv_kode = DataValidation(type="list", formula1=DROPDOWN, allow_blank=True)
dv_kode.sqref = f"A5:A500"
ws1.add_data_validation(dv_kode)

# Legend
ws1.row_dimensions[row+1].height = 18
ws1.merge_cells(f"A{row+1}:H{row+1}")
c = ws1.cell(row=row+1, column=1, value="Warna BIRU = Input manual pengguna  |  Warna HITAM = Formula otomatis  |  Paste data baru setelah baris terakhir saham")
c.font = font(False, 8, "555555", italic=True)
c.fill = fill(WARN_BG)
c.alignment = align("center", "center")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Price_Chart
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Price_Chart")
ws2.freeze_panes = "A4"

ws2.row_dimensions[1].height = 30
ws2.merge_cells("A1:M1")
c = ws2["A1"]
c.value = "📊  ANALISIS OHLCV & TEKNIKAL PER SAHAM"
c.font = font(True, 14, NAVY)
c.fill = fill(LTBLUE)
c.alignment = align("center", "center")

# Stock selector
ws2.row_dimensions[2].height = 24
ws2["A2"] = "Pilih Saham:"
ws2["A2"].font = font(True, 10, NAVY)
ws2["B2"] = STOCKS[0]
ws2["B2"].font = font(True, 11, "00008B")
ws2["B2"].fill = fill(INPUT_BG)
ws2["B2"].alignment = align("center", "center")
ws2["B2"].border = thin_border()

dv_sel = DataValidation(type="list", formula1=DROPDOWN, allow_blank=False)
dv_sel.sqref = "B2"
ws2.add_data_validation(dv_sel)

ws2["C2"] = "← Pilih kode dari dropdown"
ws2["C2"].font = font(False, 9, "888888", italic=True)

# Stats section header
header_row(ws2, row=4,
    cols=  [1,       2,       3,        4,        5,       6,       7,       8],
    labels=["Kode","Tgl Terakhir","Close","High 10H","Low 10H","Vol","Vol Avg 10H","Vol Mult"],
    bg=NAVY, fs=9)

# Stats formulas (pulling from GF_Import for selected stock)
ws2.row_dimensions[5].height = 20
stats_formulas = [
    '=B2',
    '=IFERROR(MAXIFS(GF_Import!B:B,GF_Import!A:A,B2),"N/A")',
    '=IFERROR(INDEX(GF_Import!F:F,MATCH(B2&MAXIFS(GF_Import!B:B,GF_Import!A:A,B2),GF_Import!A:A&GF_Import!B:B,0)),"N/A")',
    '=IFERROR(MAXIFS(GF_Import!D:D,GF_Import!A:A,B2),"N/A")',
    '=IFERROR(MINIFS(GF_Import!E:E,GF_Import!A:A,B2),"N/A")',
    '=IFERROR(SUMIF(GF_Import!A:A,B2,GF_Import!G:G)/COUNTIF(GF_Import!A:A,B2),"N/A")',
    '=IFERROR(SUMIF(GF_Import!A:A,B2,GF_Import!G:G)/COUNTIF(GF_Import!A:A,B2),"N/A")',
    '=IFERROR(C5/G5,"N/A")',
]
num_fmts = [None, None, '#,##0', '#,##0', '#,##0', '#,##0', '#,##0', '0.00"x"']
for col, (formula, nf) in enumerate(zip(stats_formulas, num_fmts), 1):
    c = ws2.cell(row=5, column=col, value=formula)
    c.font = font(False, 10, "000000")
    c.fill = fill(CALC_BG)
    c.alignment = align("center", "center")
    c.border = thin_border()
    if nf: c.number_format = nf

# ── Build chart data from filtered OHLCV ─────────────────────────────────────
# Extract up to 20 rows of Close + Volume for selected stock (BBRI default)
chart_start_row = 8

ws2.row_dimensions[7].height = 18
ws2["A7"] = "Data Chart (auto-filter dari GF_Import untuk saham terpilih):"
ws2["A7"].font = font(True, 9, NAVY)

header_row(ws2, 8, [1,2,3,4,5,6,7], ["No","Tanggal","Open","High","Low","Close","Volume"], bg=D_GREEN, fs=9)

# Populate chart data rows with IFERROR+INDEX/MATCH for selected stock (20 rows)
# We'll use SMALL to get sorted dates for selected stock
for i in range(1, 21):
    r = chart_start_row + i
    ws2.row_dimensions[r].height = 18

    ws2.cell(row=r, column=1, value=i).font = font(False, 9, "888888")
    ws2.cell(row=r, column=1).alignment = align("center", "center")

    # Date: nth SMALL date for selected stock
    for col, src_col in [(2,"B"),(3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G")]:
        formula = (
            f'=IFERROR(INDEX(GF_Import!{src_col}:{src_col},'
            f'MATCH(AGGREGATE(15,6,(ROW(GF_Import!A$5:A$500)-ROW(GF_Import!A$5)+1)'
            f'/(GF_Import!A$5:A$500=$B$2),{i})+ROW(GF_Import!A$5)-1,'
            f'ROW(GF_Import!A$5:A$500)-ROW(GF_Import!A$5)+1+ROW(GF_Import!A$5)-1,0)),"")'
        )
        c = ws2.cell(row=r, column=col, value=formula)
        c.fill = fill(WHITE if r % 2 == 0 else LTGRAY)
        c.border = thin_border()
        c.alignment = align("center", "center")
        c.font = font(False, 9, "000000")
        if col in (3,4,5,6): c.number_format = '#,##0'
        if col == 7: c.number_format = '#,##0'

chart_data_end = chart_start_row + 20

# MA20 + MA50 placeholder columns (cols 8 & 9)
ws2.cell(row=8, column=8, value="MA 20").font = font(True, 9, WHITE)
ws2.cell(row=8, column=8).fill = fill(D_GREEN)
ws2.cell(row=8, column=8).alignment = align("center", "center")
ws2.cell(row=8, column=9, value="% Chg").font = font(True, 9, WHITE)
ws2.cell(row=8, column=9).fill = fill(D_GREEN)
ws2.cell(row=8, column=9).alignment = align("center", "center")

for i in range(1, 21):
    r = chart_start_row + i
    # MA20 (average of last 20 closes - simplified as AVERAGE formula)
    c8 = ws2.cell(row=r, column=8,
        value=f'=IF(F{r}="","",IFERROR(AVERAGEIF(GF_Import!A:A,$B$2,GF_Import!F:F),""))')
    c8.number_format = '#,##0'
    c8.fill = fill(CALC_BG)
    c8.font = font(False, 9, "000000")
    c8.border = thin_border()
    c8.alignment = align("center", "center")

    c9 = ws2.cell(row=r, column=9,
        value=f'=IF(OR(F{r}="",F{r-1}=""),"",IFERROR((F{r}-F{r-1})/F{r-1},""))')
    c9.number_format = "0.00%"
    c9.fill = fill(CALC_BG)
    c9.font = font(False, 9, "000000")
    c9.border = thin_border()
    c9.alignment = align("center", "center")

# ── Price Line Chart ──────────────────────────────────────────────────────────
price_chart = LineChart()
price_chart.title = "Harga Close (Rp)"
price_chart.style = 10
price_chart.height = 12
price_chart.width = 22
price_chart.grouping = "standard"
price_chart.legend.position = "b"

close_ref  = Reference(ws2, min_col=6, min_row=chart_start_row+1, max_row=chart_data_end)
ma_ref     = Reference(ws2, min_col=8, min_row=chart_start_row+1, max_row=chart_data_end)
dates_ref  = Reference(ws2, min_col=2, min_row=chart_start_row+1, max_row=chart_data_end)

# openpyxl 3.1.x: jangan append Reference langsung ke chart/series.
# Gunakan add_data(), lalu set judul series secara manual.
price_chart.add_data(close_ref, titles_from_data=False)
price_chart.add_data(ma_ref, titles_from_data=False)
price_chart.set_categories(dates_ref)
price_chart.series[0].title = SeriesLabel(v="Close")
price_chart.series[0].graphicalProperties.line.solidFill = "2E5E8A"
price_chart.series[0].graphicalProperties.line.width = 18000
if len(price_chart.series) > 1:
    price_chart.series[1].title = SeriesLabel(v="MA20")
    price_chart.series[1].graphicalProperties.line.solidFill = "F59E0B"
    price_chart.series[1].graphicalProperties.line.width = 12000

ws2.add_chart(price_chart, "A30")

# ── Volume Bar Chart ─────────────────────────────────────────────────────────
vol_chart = BarChart()
vol_chart.title = "Volume"
vol_chart.style = 10
vol_chart.height = 8
vol_chart.width = 22
vol_chart.type = "col"
vol_chart.legend = None

vol_ref = Reference(ws2, min_col=7, min_row=chart_start_row+1, max_row=chart_data_end)
vol_chart.add_data(vol_ref, titles_from_data=False)
vol_chart.set_categories(dates_ref)
vol_chart.series[0].title = SeriesLabel(v="Volume")
vol_chart.series[0].graphicalProperties.solidFill = "22C55E"

ws2.add_chart(vol_chart, "A50")

set_w(ws2, {1:5, 2:14, 3:12, 4:12, 5:12, 6:12, 7:16, 8:12, 9:12})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PE_PBV_Band
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("PE_PBV_Band")
ws3.freeze_panes = "B5"

ws3.row_dimensions[1].height = 30
ws3.merge_cells("A1:P1")
c = ws3["A1"]
c.value = "💰  ANALISIS VALUASI — PE BAND & PBV BAND"
c.font = font(True, 14, NAVY)
c.fill = fill(LTBLUE)
c.alignment = align("center", "center")

# Header labels rows
ws3.row_dimensions[2].height = 14
ws3.merge_cells("B2:F2")
ws3["B2"] = "INPUT DATA"
ws3["B2"].font = font(True, 9, WHITE)
ws3["B2"].fill = fill(NAVY)
ws3["B2"].alignment = align("center", "center")

ws3.merge_cells("G2:K2")
ws3["G2"] = "PE BAND (input historis — lihat RTI/BMAD)"
ws3["G2"].font = font(True, 9, WHITE)
ws3["G2"].fill = fill(D_GREEN)
ws3["G2"].alignment = align("center", "center")

ws3.merge_cells("L2:P2")
ws3["L2"] = "PBV BAND (input historis)"
ws3["L2"].font = font(True, 9, WHITE)
ws3["L2"].fill = fill(D_RED)
ws3["L2"].alignment = align("center", "center")

# Column headers row 3
header_row(ws3, 3,
    cols=  [1,      2,       3,    4,    5,      6,       7,      8,       9,      10,     11,     12,     13,      14,     15,      16],
    labels=["Kode","Harga","EPS\n(TTM)","BVPS","PE\nSaat ini","PBV\nSaat ini",
            "PE\nMin","PE\n-1SD","PE\nMedian","PE\n+1SD","PE\nMax",
            "PBV\nMin","PBV\n-1SD","PBV\nMedian","PBV\n+1SD","PBV\nMax"],
    bg=NAVY, fs=9, height=34)
ws3.row_dimensions[3].height = 34

# Row 4 header: OUTPUT
header_row(ws3, 4,
    cols=  [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16],
    labels=["","","","","PE Curr","PBV Curr",
            "Impl. PE Min","Impl. PE -1SD","Impl. PE Med","Impl. PE +1SD","Impl. PE Max",
            "Impl.PBV Min","Impl.PBV -1SD","Impl.PBV Med","Impl.PBV +1SD","Impl.PBV Max"],
    bg=HDR_AMBER if False else "3D5A80", fs=8, height=20)

# Actually, let me redo: rows 3 = input headers, then rows 5+ per stock (input + output separated)
# Simpler layout: one block per stock (inputs row, outputs row, label row)

ws3.delete_rows(3, 2)  # remove those headers, redo cleanly

# Clean headers
header_row(ws3, 3,
    cols=  [1,    2,       3,    4,    5,      6,      7,      8,       9,      10,      11],
    labels=["KODE","Harga\nClose","EPS\nTTM","BVPS","PE\nMin","PE\n-1SD","PE\nMedian","PE\n+1SD","PE\nMax",
            "PBV\nMin","PBV\nMax"],
    bg=NAVY, fs=9, height=30)

# Add more PE/PBV cols
header_row(ws3, 3,
    cols=  [10,11,12,13,14],
    labels=["PBV\nMin","PBV\n-1SD","PBV\nMedian","PBV\n+1SD","PBV\nMax"],
    bg=D_RED, fs=9, height=30)

# Output header (row 4)
header_row(ws3, 4,
    cols=  [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18],
    labels=["","","PE Curr","PBV Curr",
            "Impl Min","Impl -1SD","Impl Med","Impl +1SD","Impl Max",
            "Impl Min","Impl -1SD","Impl Med","Impl +1SD","Impl Max",
            "PE %ile","PBV %ile","PE Label","PBV Label"],
    bg="3D5A80", fs=8, height=16)
ws3["A4"] = "↑ INPUT"
ws3["A4"].font = font(True, 8, WHITE)
ws3["A4"].fill = fill("3D5A80")
ws3["A4"].alignment = align("center", "center")

# Stock input rows
stock_defaults = {
    "BBRI": (4870, 520,  2800, 8,  10, 13, 16, 20, 1.2, 1.5, 1.9, 2.3, 3.0),
    "BBCA": (9550, 690,  4200, 20, 24, 28, 32, 38, 3.5, 4.2, 5.0, 5.8, 7.0),
    "ADRO": (2420, 450,  1800, 3,  5,  7,  9,  12, 0.8, 1.1, 1.4, 1.8, 2.5),
    "TLKM": (3220, 200,  1500, 10, 13, 16, 19, 24, 1.5, 2.0, 2.5, 3.0, 3.8),
    "GOTO": (70,   -5,   50,   0,  0,  0,  0,  0,  1.0, 1.5, 2.0, 2.8, 4.0),
}

EMPTY_PE = (None,) * 13   # saham tanpa data → input kosong, user isi manual
row = 5
stock_row_map = {}
for kode in STOCKS:
    # Prioritas: data live dari scraper → contoh hardcoded → kosong
    pe_vals = pe_row_from_info(get_stock_info(kode)) or stock_defaults.get(kode) or EMPTY_PE
    (price, eps, bvps, pe_min, pe_1sd, pe_med, pe_p1sd, pe_max,
     pbv_min, pbv_1sd, pbv_med, pbv_p1sd, pbv_max) = pe_vals
    stock_row_map[kode] = row
    ws3.row_dimensions[row].height = 22

    # Input cells (blue font)
    inputs = [kode, price, eps, bvps,
              pe_min, pe_1sd, pe_med, pe_p1sd, pe_max,
              pbv_min, pbv_1sd, pbv_med, pbv_p1sd, pbv_max]
    for col, val in enumerate(inputs, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = font(False, 10, "00008B")
        c.fill = fill(INPUT_BG)
        c.border = thin_border()
        c.alignment = align("center" if col > 1 else "left", "center")
        if col in (2,3,4):   c.number_format = "#,##0.00"
        if col in range(5,15): c.number_format = "0.0"

    # Derived: Current PE (col 15)
    c15 = ws3.cell(row=row, column=15,
        value=f'=IFERROR(B{row}/C{row},"-")')
    c15.font = font(False, 10, "000000")
    c15.fill = fill(CALC_BG)
    c15.number_format = "0.0"
    c15.border = thin_border()
    c15.alignment = align("center", "center")

    # Current PBV (col 16)
    c16 = ws3.cell(row=row, column=16,
        value=f'=IFERROR(B{row}/D{row},"-")')
    c16.font = font(False, 10, "000000")
    c16.fill = fill(CALC_BG)
    c16.number_format = "0.0"
    c16.border = thin_border()
    c16.alignment = align("center", "center")

    # Implied prices at PE bands (cols 17–21)
    pe_cols = [5, 6, 7, 8, 9]
    for i, (out_col, pe_col) in enumerate(zip(range(17, 22), pe_cols)):
        c = ws3.cell(row=row, column=out_col,
            value=f'=IFERROR({get_column_letter(pe_col)}{row}*C{row},"-")')
        c.font = font(False, 10, "000000")
        c.fill = fill(CALC_BG)
        c.number_format = "#,##0"
        c.border = thin_border()
        c.alignment = align("center", "center")

    # Implied prices at PBV bands (cols 22–26)
    pbv_cols = [10, 11, 12, 13, 14]
    for out_col, pbv_col in zip(range(22, 27), pbv_cols):
        c = ws3.cell(row=row, column=out_col,
            value=f'=IFERROR({get_column_letter(pbv_col)}{row}*D{row},"-")')
        c.font = font(False, 10, "000000")
        c.fill = fill(CALC_BG)
        c.number_format = "#,##0"
        c.border = thin_border()
        c.alignment = align("center", "center")

    # PE percentile (col 27)
    c27 = ws3.cell(row=row, column=27,
        value=f'=IFERROR((O{row}-E{row})/(I{row}-E{row})*100,"-")')
    c27.font = font(False, 10, "000000")
    c27.fill = fill(CALC_BG)
    c27.number_format = "0.0"
    c27.border = thin_border()
    c27.alignment = align("center", "center")

    # PBV percentile (col 28)
    c28 = ws3.cell(row=row, column=28,
        value=f'=IFERROR((P{row}-J{row})/(N{row}-J{row})*100,"-")')
    c28.font = font(False, 10, "000000")
    c28.fill = fill(CALC_BG)
    c28.number_format = "0.0"
    c28.border = thin_border()
    c28.alignment = align("center", "center")

    # PE Valuation Label (col 29)
    ws3.cell(row=row, column=29,
        value=(f'=IFERROR(IF(AA{row}<15,"🟢 Deep Under",IF(AA{row}<35,"🟡 Undervalued",'
               f'IF(AA{row}<65,"⚪ Fair Value",IF(AA{row}<85,"🟠 Overvalued","🔴 Deep Over")))),"N/A")'))
    ws3.cell(row=row, column=29).font = font(True, 10, "000000")
    ws3.cell(row=row, column=29).fill = fill(CALC_BG)
    ws3.cell(row=row, column=29).border = thin_border()
    ws3.cell(row=row, column=29).alignment = align("center", "center")

    # PBV Valuation Label (col 30)
    ws3.cell(row=row, column=30,
        value=(f'=IFERROR(IF(AB{row}<15,"🟢 Deep Under",IF(AB{row}<35,"🟡 Undervalued",'
               f'IF(AB{row}<65,"⚪ Fair Value",IF(AB{row}<85,"🟠 Overvalued","🔴 Deep Over")))),"N/A")'))
    ws3.cell(row=row, column=30).font = font(True, 10, "000000")
    ws3.cell(row=row, column=30).fill = fill(CALC_BG)
    ws3.cell(row=row, column=30).border = thin_border()
    ws3.cell(row=row, column=30).alignment = align("center", "center")

    row += 1

# Header labels for output cols
output_headers = {
    15: "PE Curr", 16: "PBV Curr",
    17: "Impl\nPE Min", 18: "Impl\nPE -1SD", 19: "Impl\nPE Med",
    20: "Impl\nPE +1SD", 21: "Impl\nPE Max",
    22: "Impl\nPBV Min", 23: "Impl\nPBV -1SD", 24: "Impl\nPBV Med",
    25: "Impl\nPBV +1SD", 26: "Impl\nPBV Max",
    27: "PE\n%ile", 28: "PBV\n%ile", 29: "PE Label", 30: "PBV Label"
}
ws3.row_dimensions[3].height = 34
for col, lbl in output_headers.items():
    c = ws3.cell(row=3, column=col, value=lbl)
    c.font = font(True, 8, WHITE)
    c.fill = fill("2E5E8A" if col <= 21 else ("8B1A1A" if col <= 26 else "444466"))
    c.alignment = align("center", "center", wrap=True)
    c.border = thin_border("888888")

# Upside/Downside table (cols 31–40 = upside% to each band level)
upside_label_row = 2
ws3.cell(row=upside_label_row, column=31, value="UPSIDE / DOWNSIDE KE SETIAP BAND LEVEL")
ws3.cell(row=upside_label_row, column=31).font = font(True, 10, WHITE)
ws3.cell(row=upside_label_row, column=31).fill = fill(D_GREEN)
ws3.merge_cells(f"AE{upside_label_row}:AJ{upside_label_row}")
ws3.cell(row=upside_label_row, column=31).alignment = align("center", "center")

up_headers = ["Ke PE Min","Ke PE -1SD","Ke PE Med","Ke PE +1SD","Ke PE Max",
              "Ke PBV Min","Ke PBV -1SD","Ke PBV Med","Ke PBV +1SD","Ke PBV Max"]
for i, lbl in enumerate(up_headers, 31):
    c = ws3.cell(row=3, column=i, value=lbl)
    c.font = font(True, 8, WHITE)
    c.fill = fill(D_GREEN if i <= 35 else "8B4513")
    c.alignment = align("center", "center", wrap=True)
    c.border = thin_border("888888")

for r in range(5, 5+len(STOCKS)):
    for i, out_col in enumerate(range(17, 27)):
        upcol = 31 + i
        c = ws3.cell(row=r, column=upcol,
            value=f'=IFERROR(({get_column_letter(out_col)}{r}-B{r})/B{r},"-")')
        c.font = font(False, 9, "000000")
        c.fill = fill(CALC_BG)
        c.number_format = "+0.0%;-0.0%;-"
        c.border = thin_border()
        c.alignment = align("center", "center")

# Column widths sheet3
set_w(ws3, {1:8, 2:10, 3:10, 4:10, 5:9, 6:9, 7:9, 8:9, 9:9,
            10:9, 11:9, 12:9, 13:9, 14:9, 15:9, 16:9,
            17:10, 18:10, 19:10, 20:10, 21:10,
            22:10, 23:10, 24:10, 25:10, 26:10,
            27:9, 28:9, 29:14, 30:14,
            31:12, 32:12, 33:12, 34:12, 35:12,
            36:12, 37:12, 38:12, 39:12, 40:12})

# ── PE Band Bar Chart (for BBRI as example) ───────────────────────────────────
# Create a small helper table for chart
chart_help_row = 12
ws3.cell(row=chart_help_row, column=1, value=f"PE Band Chart Data ({STOCKS[0]})")
ws3.cell(row=chart_help_row, column=1).font = font(True, 9, NAVY)

ws3.cell(row=chart_help_row+1, column=1, value="Level")
ws3.cell(row=chart_help_row+1, column=2, value="Implied Price")
ws3.cell(row=chart_help_row+1, column=3, value="Harga Saat Ini")
levels = ["Min", "-1SD", "Median", "+1SD", "Max"]
for i, lbl in enumerate(levels):
    ws3.cell(row=chart_help_row+2+i, column=1, value=lbl)
    ws3.cell(row=chart_help_row+2+i, column=2, value=f"=Q5")
    ws3.cell(row=chart_help_row+2+i, column=3, value=f"=B5")
for i, col_offset in enumerate([17,18,19,20,21]):
    ws3.cell(row=chart_help_row+2+i, column=2, value=f"={get_column_letter(col_offset)}5")

pe_band_chart = BarChart()
pe_band_chart.type = "bar"
pe_band_chart.title = f"PE Band Implied Price vs Harga Saat Ini ({STOCKS[0]})"
pe_band_chart.style = 10
pe_band_chart.height = 10
pe_band_chart.width = 18
pe_band_chart.grouping = "clustered"

data_ref   = Reference(ws3, min_col=2, min_row=chart_help_row+1, max_row=chart_help_row+6)
curr_ref   = Reference(ws3, min_col=3, min_row=chart_help_row+1, max_row=chart_help_row+6)
cats_ref   = Reference(ws3, min_col=1, min_row=chart_help_row+2, max_row=chart_help_row+6)

pe_band_chart.add_data(data_ref, titles_from_data=True)
pe_band_chart.add_data(curr_ref, titles_from_data=True)
pe_band_chart.set_categories(cats_ref)
pe_band_chart.series[0].title = SeriesLabel(v="Implied Price")
pe_band_chart.series[0].graphicalProperties.solidFill = "2E5E8A"
if len(pe_band_chart.series) > 1:
    pe_band_chart.series[1].title = SeriesLabel(v="Harga Saat Ini")
    pe_band_chart.series[1].graphicalProperties.solidFill = "F59E0B"

ws3.add_chart(pe_band_chart, "A20")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Consensus_EPS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Consensus_EPS")
ws4.freeze_panes = "B4"

ws4.row_dimensions[1].height = 30
ws4.merge_cells("A1:R1")
c = ws4["A1"]
c.value = "📋  CONSENSUS ESTIMATE & ANALYST RATING"
c.font = font(True, 14, NAVY)
c.fill = fill(LTBLUE)
c.alignment = align("center", "center")

# Section headers row 2
for merge, bg, text in [
    ("B2:D2", D_GREEN,  "ANALYST RATING"),
    ("E2:G2", "1B3A59", "TARGET PRICE"),
    ("H2:J2", "3A1B59", "EPS ESTIMATE"),
    ("K2:M2", "1B4A2A", "REVENUE (Rp Miliar)"),
    ("N2:O2", "5A3A1B", "CONSENSUS OUTPUT"),
    ("P2:R2", "2A4A3A", "SKOR & LABEL"),
]:
    ws4.merge_cells(merge)
    col = int(merge[1]) if merge[1].isdigit() else ord(merge[0])-64
    # Get column from merge string
    start_col = merge.split(":")[0]
    col_num = 0
    for char in start_col:
        if char.isalpha():
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
    c = ws4.cell(row=2, column=col_num, value=text)
    c.font = font(True, 9, WHITE)
    c.fill = fill(bg)
    c.alignment = align("center", "center")
    ws4.row_dimensions[2].height = 18

# Column headers row 3
col3_headers = [
    "KODE",
    "Buy","Hold","Sell",
    "TP Low","TP Mean","TP High",
    "EPS FY0","EPS FY1","EPS Grwth",
    "Rev FY0","Rev FY1","Rev Grwth",
    "Next Earnings","Catalyst",
    "Consensus\nScore","Consensus\nLabel","Upside\nke Mean"
]
header_row(ws4, 3, list(range(1, len(col3_headers)+1)), col3_headers, bg=NAVY, fs=9, height=30)

# Data rows
consensus_data = {
    "BBRI": (18, 5, 1, 4500, 5400, 6200, 520, 610, None, 145000, 162000, None, "2026-07-28", "Dividen Final"),
    "BBCA": (22, 3, 0, 9000, 10200, 11500, 690, 780, None, 200000, 225000, None, "2026-07-30", "Rights Issue"),
    "ADRO": (12, 8, 2, 2200, 2800, 3400, 450, 500, None, 85000, 90000,  None, "2026-08-05", "Harga Batubara"),
    "TLKM": (10, 6, 4, 2800, 3400, 3900, 200, 220, None, 140000, 155000, None, "2026-08-15", "5G Expansion"),
    "GOTO": (8,  10, 5, 55,  85,   120,  -5,  2,   None, 18000,  25000,  None, "2026-08-20", "GTV Growth"),
}

EMPTY_CONS = (None,) * 14   # saham tanpa data → input kosong, user isi manual
row = 4
cons_rows = {}
for kode in STOCKS:
    # Prioritas: data live dari scraper → contoh hardcoded → kosong
    cons_vals = cons_row_from_info(get_stock_info(kode)) or consensus_data.get(kode) or EMPTY_CONS
    (buy, hold, sell, tp_lo, tp_mn, tp_hi,
     eps0, eps1, _, rev0, rev1, __, earn_dt, catalyst) = cons_vals
    cons_rows[kode] = row
    ws4.row_dimensions[row].height = 22

    # Input cells
    vals = [kode, buy, hold, sell, tp_lo, tp_mn, tp_hi, eps0, eps1, None, rev0, rev1, None, earn_dt, catalyst]
    for col, val in enumerate(vals, 1):
        c = ws4.cell(row=row, column=col, value=val)
        c.font = font(False, 10, "00008B" if val is not None and col != 1 else "000000" if col == 1 else "888888")
        c.fill = fill(INPUT_BG if (val is not None and col != 1) else (WHITE if col == 1 else CALC_BG))
        c.border = thin_border()
        c.alignment = align("center" if col > 1 else "left", "center")
        if col in (5,6,7):   c.number_format = "#,##0"
        if col in (8,9):     c.number_format = "#,##0.00"
        if col in (11,12):   c.number_format = "#,##0"

    # EPS Growth formula (col 10)
    eps_grwth = ws4.cell(row=row, column=10,
        value=f'=IFERROR((I{row}-H{row})/ABS(H{row}),"-")')
    eps_grwth.font = font(False, 10, "000000")
    eps_grwth.fill = fill(CALC_BG)
    eps_grwth.number_format = "0.0%"
    eps_grwth.border = thin_border()
    eps_grwth.alignment = align("center", "center")

    # Revenue Growth formula (col 13)
    rev_grwth = ws4.cell(row=row, column=13,
        value=f'=IFERROR((L{row}-K{row})/K{row},"-")')
    rev_grwth.font = font(False, 10, "000000")
    rev_grwth.fill = fill(CALC_BG)
    rev_grwth.number_format = "0.0%"
    rev_grwth.border = thin_border()
    rev_grwth.alignment = align("center", "center")

    # Consensus Score (col 16): (Buy*100 + Hold*50 + Sell*0) / total
    cons_score = ws4.cell(row=row, column=16,
        value=f'=IFERROR((B{row}*100+C{row}*50)/(B{row}+C{row}+D{row}),"-")')
    cons_score.font = font(True, 10, "000000")
    cons_score.fill = fill(CALC_BG)
    cons_score.number_format = "0.0"
    cons_score.border = thin_border()
    cons_score.alignment = align("center", "center")

    # Consensus Label (col 17)
    cons_label = ws4.cell(row=row, column=17,
        value=(f'=IFERROR(IF(P{row}>=70,"🟢 BULLISH",IF(P{row}>=40,"🟡 NEUTRAL","🔴 BEARISH")),"N/A")'))
    cons_label.font = font(True, 10, "000000")
    cons_label.fill = fill(CALC_BG)
    cons_label.border = thin_border()
    cons_label.alignment = align("center", "center")

    # Upside ke TP Mean (col 18) - needs current price from PE_PBV_Band
    # Using VLOOKUP to get price from sheet 3
    upside_mean = ws4.cell(row=row, column=18,
        value=f'=IFERROR((F{row}-VLOOKUP(A{row},PE_PBV_Band!A:B,2,0))/VLOOKUP(A{row},PE_PBV_Band!A:B,2,0),"-")')
    upside_mean.font = font(False, 10, "000000")
    upside_mean.fill = fill(CALC_BG)
    upside_mean.number_format = "+0.0%;-0.0%;-"
    upside_mean.border = thin_border()
    upside_mean.alignment = align("center", "center")

    row += 1

set_w(ws4, {i: w for i, w in enumerate(
    [8, 7, 7, 7, 10, 10, 10, 11, 11, 9, 12, 12, 9, 14, 15, 11, 12, 11], 1)})

# ── Consensus Pie Chart (for BBRI) ─────────────────────────────────────────
pie_help_row = 11
ws4.cell(row=pie_help_row, column=1, value=f"Pie Data ({STOCKS[0]})")
ws4.cell(row=pie_help_row+1, column=1, value="Buy")
ws4.cell(row=pie_help_row+2, column=1, value="Hold")
ws4.cell(row=pie_help_row+3, column=1, value="Sell")
ws4.cell(row=pie_help_row+1, column=2, value="=B4")
ws4.cell(row=pie_help_row+2, column=2, value="=C4")
ws4.cell(row=pie_help_row+3, column=2, value="=D4")

pie = PieChart()
pie.title = f"Distribusi Rating Analis ({STOCKS[0]})"
pie.style = 10
pie.height = 10
pie.width = 14

pie_data = Reference(ws4, min_col=2, min_row=pie_help_row+1, max_row=pie_help_row+3)
pie_cats = Reference(ws4, min_col=1, min_row=pie_help_row+1, max_row=pie_help_row+3)
pie.add_data(pie_data, titles_from_data=False)
pie.dataLabels = None
pie.set_categories(pie_cats)

ws4.add_chart(pie, "A18")

# TP Range Bar Chart
tp_help_row = pie_help_row
ws4.cell(row=tp_help_row, column=4, value="TP Range")
ws4.cell(row=tp_help_row+1, column=4, value="Low")
ws4.cell(row=tp_help_row+2, column=4, value="Mean")
ws4.cell(row=tp_help_row+3, column=4, value="High")
ws4.cell(row=tp_help_row+4, column=4, value="Harga Saat Ini")

for i, src_col in enumerate([5, 6, 7], 1):
    ws4.cell(row=tp_help_row+i, column=5,
        value=f"={get_column_letter(src_col)}4").number_format = "#,##0"

ws4.cell(row=tp_help_row+4, column=5,
    value="=VLOOKUP(A4,PE_PBV_Band!A:B,2,0)").number_format = "#,##0"

tp_chart = BarChart()
tp_chart.type = "col"
tp_chart.title = f"TP Range vs Harga ({STOCKS[0]})"
tp_chart.style = 10
tp_chart.height = 10
tp_chart.width = 14
tp_data = Reference(ws4, min_col=5, min_row=tp_help_row+1, max_row=tp_help_row+4)
tp_cats = Reference(ws4, min_col=4, min_row=tp_help_row+1, max_row=tp_help_row+4)
tp_chart.add_data(tp_data, titles_from_data=False)
tp_chart.set_categories(tp_cats)
tp_chart.series[0].graphicalProperties.solidFill = "2E5E8A"

ws4.add_chart(tp_chart, "P18")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Brief_Sync
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Brief_Sync")

ws5.row_dimensions[1].height = 30
ws5.merge_cells("A1:F1")
c = ws5["A1"]
c.value = "🔄  BRIEF SYNC — Output untuk Scalping Brief Generator App"
c.font = font(True, 14, NAVY)
c.fill = fill(LTBLUE)
c.alignment = align("center", "center")

ws5.row_dimensions[2].height = 20
ws5.merge_cells("A2:F2")
c = ws5["A2"]
c.value = "Copy nilai dari kolom NILAI ke input field yang sesuai di Scalping Brief Generator App"
c.font = font(False, 9, "555555", italic=True)
c.fill = fill(WARN_BG)
c.alignment = align("center", "center")

row = 4
for kode in STOCKS:
    pe_r = 5 + list(STOCKS).index(kode)
    cs_r = 4 + list(STOCKS).index(kode)

    # Section divider
    ws5.row_dimensions[row].height = 24
    ws5.merge_cells(f"A{row}:F{row}")
    c = ws5.cell(row=row, column=1, value=f"▶  {kode}  — SCALPING BRIEF DATA")
    c.font = font(True, 11, WHITE)
    c.fill = fill(NAVY)
    c.alignment = align("left", "center")
    row += 1

    # Table header
    header_row(ws5, row, [1,2,3,4,5,6],
        ["MODULE","FIELD","NILAI","SATUAN","CATATAN","STATUS"],
        bg=BLUE, fs=9, height=18)
    row += 1

    sync_rows = [
        # Module, Field, Formula, Unit, Note
        ("Kandidat","Kode Saham",   f"=PE_PBV_Band!A{pe_r}",           "–",     "Kode IDX"),
        ("Kandidat","Harga Close",  f"=PE_PBV_Band!B{pe_r}",           "Rp",    "Dari GF_Import"),
        ("Valuasi","PE Saat Ini",   f"=PE_PBV_Band!O{pe_r}",           "x",     "EPS TTM"),
        ("Valuasi","PBV Saat Ini",  f"=PE_PBV_Band!P{pe_r}",           "x",     "BVPS"),
        ("Valuasi","PE Implied Med",f"=PE_PBV_Band!S{pe_r}",           "Rp",    "PE Median Band"),
        ("Valuasi","Upside PE Med", f"=PE_PBV_Band!AG{pe_r}",          "%",     "Ke PE Median"),
        ("Valuasi","Label Valuasi", f"=PE_PBV_Band!AC{pe_r}",          "–",     "Auto dari percentile"),
        ("Consensus","Rating",      f"=Consensus_EPS!Q{cs_r}",         "–",     "Bullish/Neutral/Bear"),
        ("Consensus","Cons Score",  f"=Consensus_EPS!P{cs_r}",         "/100",  "Weighted avg"),
        ("Consensus","TP Mean",     f"=Consensus_EPS!F{cs_r}",         "Rp",    "Mean target analis"),
        ("Consensus","Upside TP",   f"=Consensus_EPS!R{cs_r}",         "%",     "Vs harga saat ini"),
        ("Consensus","EPS FY0",     f"=Consensus_EPS!H{cs_r}",         "Rp",    "TTM / FY aktual"),
        ("Consensus","EPS FY1 Est", f"=Consensus_EPS!I{cs_r}",         "Rp",    "Estimasi analis"),
        ("Consensus","EPS Growth",  f"=Consensus_EPS!J{cs_r}",         "%",     "YoY"),
        ("Consensus","Next Earnings",f"=Consensus_EPS!N{cs_r}",        "–",     "Tanggal LK berikutnya"),
    ]

    for mod, field, formula, unit, note in sync_rows:
        ws5.row_dimensions[row].height = 18

        c1 = ws5.cell(row=row, column=1, value=mod)
        c1.font = font(True, 9, WHITE)
        c1.fill = fill({"Kandidat": "1B3A59", "Valuasi": D_GREEN, "Consensus": "5A1B59"}[mod])
        c1.alignment = align("center", "center")
        c1.border = thin_border()

        c2 = ws5.cell(row=row, column=2, value=field)
        c2.font = font(False, 9, "000000")
        c2.fill = fill(LTGRAY)
        c2.alignment = align("left", "center")
        c2.border = thin_border()

        c3 = ws5.cell(row=row, column=3, value=formula)
        c3.font = font(True, 10, "000000")
        c3.fill = fill(CALC_BG)
        c3.alignment = align("center", "center")
        c3.border = thin_border()

        c4 = ws5.cell(row=row, column=4, value=unit)
        c4.font = font(False, 9, "888888")
        c4.alignment = align("center", "center")
        c4.border = thin_border()

        c5 = ws5.cell(row=row, column=5, value=note)
        c5.font = font(False, 8, "888888", italic=True)
        c5.alignment = align("left", "center")
        c5.border = thin_border()

        # Status: OK / Perlu Check
        c6 = ws5.cell(row=row, column=6, value=f'=IF(C{row}="","⚠️ Kosong","✅ Siap")')
        c6.font = font(True, 9, "000000")
        c6.fill = fill(CALC_BG)
        c6.alignment = align("center", "center")
        c6.border = thin_border()

        row += 1

    row += 1  # gap between stocks

set_w(ws5, {1:14, 2:22, 3:16, 4:8, 5:28, 6:10})

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET ORDER & COLORS
# ═══════════════════════════════════════════════════════════════════════════════
TAB_COLORS = {
    "GF_Import":    "1B3A59",
    "Price_Chart":  "2E5E8A",
    "PE_PBV_Band":  "155724",
    "Consensus_EPS":"7B1A1A",
    "Brief_Sync":   "856404",
}
for sheet_name, color in TAB_COLORS.items():
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = color

# ── SAVE ──────────────────────────────────────────────────────────────────────
# Simpan di folder yang sama dengan file script saat dijalankan secara lokal.
# Contoh: C:\Users\IHSAN\Claude\Projects\Trading\Scalping_Analysis.xlsx
script_dir = os.path.dirname(os.path.abspath(__file__))
if args.timestamp:
    nama_file = "Scalping_Analysis_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
else:
    nama_file = "Scalping_Analysis.xlsx"
output_path = os.path.join(script_dir, nama_file)

try:
    wb.save(output_path)
    print(f"Saved: {output_path}")
except PermissionError:
    print("Tutup Scalping_Analysis.xlsx di Excel dulu, lalu jalankan ulang.")
